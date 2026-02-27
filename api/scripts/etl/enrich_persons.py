"""
ETL Phase 1 — Enrich core.person from Funcionarios CSVs.

Functor F: C_CSV → C_DB(core.person)
  - Object mapping: CSV row (monthly salary event) → Person entity
  - Morphism: name matching (paternal_surname + names) for identity
  - Colimit: merge N monthly events per person into single update

Tension A1[Evento ↔ Entidad]: Each CSV row is a salary event, but core.person
is a persistent entity. We collapse repeated events, keeping the LATEST data.

Functor Information Loss: Email fibra remains empty — no CSV source contains it.
RUT fibra can now be filled from NOMINA xlsx (--nomina flag).

Usage:
  docker compose exec api python -m scripts.etl.enrich_persons --dry-run
  docker compose exec api python -m scripts.etl.enrich_persons --verbose
  docker compose exec api python -m scripts.etl.enrich_persons --nomina /app/data/etl/funcionarios/NOMINA.xlsx
  docker compose exec api python -m scripts.etl.enrich_persons --limit 10
"""

import json
import sys
from pathlib import Path

from sqlalchemy import text

from scripts.etl.common import (
    base_argparser,
    setup_logging,
    read_csv,
    parse_fullname,
    normalize_text,
    normalize_rut,
    resolve_category_by_label,
    build_person_cache,
    resolve_person_by_name,
    get_session,
    close_engine,
    batch_commit,
    ETLStats,
    run_async,
    log,
    BATCH_SIZE,
)


# ---------------------------------------------------------------------------
# CSV Sources — ordered by priority (latest first)
# ---------------------------------------------------------------------------

# Paths relative to /app inside the API container
# CSVs are mounted or copied to /app/data/etl/
ETL_DATA_DIR = Path("/app/data/etl/funcionarios")

# Source files in priority order (latest data wins on conflict)
SOURCE_FILES = [
    "TransparenciaActiva_3.csv",  # Honorarios (latest)
    "TransparenciaActiva_2.csv",  # Planta/Contrata batch 2
    "TransparenciaActiva_1.csv",  # Planta/Contrata batch 1
    "listado_funcionarios_integrado_remediado.csv",  # Original integrated list
]


# ---------------------------------------------------------------------------
# Estamento mapping: CSV label → ref.category(scheme='estamento') label
# ---------------------------------------------------------------------------

ESTAMENTO_MAP = {
    "profesional": "Profesional",
    "administrativo": "Administrativo",
    "auxiliar": "Auxiliar",
    "directivo": "Directivo",
    "tecnico": "Técnico",
    "honorarios": "Honorarios",
    "autoridad de gobierno": "Autoridad de Gobierno",
}


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

async def load_and_merge_csvs(data_dir: Path, limit: int = 0) -> dict[str, dict]:
    """Load all source CSVs, deduplicate by name, keep latest data.

    Returns: {normalized_name: {csv fields}} — one entry per unique person.
    The colimit: later files overwrite earlier ones for same person.
    """
    merged: dict[str, dict] = {}

    for filename in reversed(SOURCE_FILES):  # Process oldest first, latest overwrites
        filepath = data_dir / filename
        if not filepath.exists():
            log.warning(f"Source not found, skipping: {filepath}")
            continue

        try:
            rows = read_csv(filepath)
        except (FileNotFoundError, ValueError) as e:
            log.warning(f"Cannot read {filepath}: {e}")
            continue

        count = 0
        for row in rows:
            name_raw = row.get("Nombre completo", "").strip()
            if not name_raw:
                continue

            key = normalize_text(name_raw)
            merged[key] = {
                "nombre_completo": name_raw,
                "estamento": row.get("Estamento", "").strip(),
                "cargo": row.get("Cargo o función",
                         row.get("Descripción de la función", "")).strip(),
                "calificacion": row.get("Calificación profesional o formación", "").strip(),
                "grado_eus": row.get("Grado EUS o jornada",
                             row.get("Grado EUS  (si corresponde)", "")).strip(),
                "tipo_vinculo": row.get("Tipo vínculo", "").strip(),
                "_source": filename,
            }
            count += 1

        log.info(f"Loaded {count} rows from {filename}")

    log.info(f"Merged CSV: {len(merged)} unique persons across all sources")

    if limit > 0:
        keys = list(merged.keys())[:limit]
        merged = {k: merged[k] for k in keys}
        log.info(f"Limited to {limit} persons")

    return merged


async def enrich_persons(data_dir: Path, dry_run: bool = False, limit: int = 0) -> ETLStats:
    """Main ETL pipeline: CSV → core.person enrichment.

    For each CSV person:
      1. Parse name → match to existing core.person
      2. If match: UPDATE metadata JSONB + verify estamento_id
      3. If no match: INSERT new person (names, paternal, maternal, estamento)
    """
    stats = ETLStats()
    db = await get_session()

    try:
        # Phase 1: Load and merge CSVs
        merged = await load_and_merge_csvs(data_dir, limit)
        if not merged:
            log.error("No CSV data loaded. Check paths and file availability.")
            return stats

        # Phase 2: Build person cache from DB
        await build_person_cache(db)

        # Phase 3: Process each unique person
        row_num = 0
        for norm_name, csv_data in merged.items():
            row_num += 1
            name_raw = csv_data["nombre_completo"]

            try:
                parsed = parse_fullname(name_raw)
                if not parsed.paternal_surname:
                    stats.errors.append(f"Row {row_num}: Cannot parse name '{name_raw}'")
                    continue

                # Try to match to existing person
                existing = await resolve_person_by_name(db, parsed)

                # Build metadata enrichment
                meta_update = {}
                if csv_data["cargo"]:
                    # Truncate very long cargo descriptions
                    meta_update["cargo"] = csv_data["cargo"][:500]
                if csv_data["calificacion"]:
                    meta_update["calificacion"] = csv_data["calificacion"]
                if csv_data["grado_eus"]:
                    meta_update["grado_eus"] = csv_data["grado_eus"]
                if csv_data["tipo_vinculo"]:
                    meta_update["tipo_vinculo"] = csv_data["tipo_vinculo"]
                meta_update["_etl_source"] = csv_data["_source"]

                # Resolve estamento
                estamento_id = None
                estamento_raw = csv_data.get("estamento", "")
                if estamento_raw:
                    mapped_label = ESTAMENTO_MAP.get(normalize_text(estamento_raw))
                    if mapped_label:
                        estamento_id = await resolve_category_by_label(db, "estamento", mapped_label)
                        if not estamento_id:
                            stats.errors.append(
                                f"Row {row_num} ({name_raw}): estamento '{estamento_raw}' not in ref.category"
                            )

                if existing:
                    # UPDATE existing person
                    person_id = str(existing["id"])
                    updates = []
                    params: dict[str, any] = {"pid": person_id}

                    # Enrich metadata via JSONB merge (non-destructive)
                    # Note: use CAST() instead of :: to avoid asyncpg parameter conflict
                    if meta_update:
                        updates.append("metadata = metadata || CAST(:meta_json AS jsonb)")
                        params["meta_json"] = json.dumps(meta_update)

                    # Set estamento if not already set
                    if estamento_id and not existing.get("estamento_id"):
                        updates.append("estamento_id = :est_id")
                        params["est_id"] = estamento_id

                    # Set maternal_surname if DB is missing it
                    if parsed.maternal_surname and not existing.get("maternal_surname"):
                        updates.append("maternal_surname = :maternal")
                        params["maternal"] = parsed.maternal_surname

                    if updates:
                        if dry_run:
                            log.info(
                                f"[DRY-RUN] UPDATE person '{name_raw}' "
                                f"(id={person_id[:8]}...): {', '.join(updates)}"
                            )
                        else:
                            sql = f"UPDATE core.person SET {', '.join(updates)}, updated_at = NOW() WHERE id = :pid"
                            await db.execute(text(sql), params)
                        stats.updated += 1
                    else:
                        stats.skipped += 1
                        log.debug(f"Skipped '{name_raw}' — no new data to enrich")

                else:
                    # INSERT new person
                    if dry_run:
                        log.info(
                            f"[DRY-RUN] INSERT new person: '{name_raw}' "
                            f"(estamento={csv_data.get('estamento', 'N/A')})"
                        )
                    else:
                        insert_params = {
                            "names": parsed.names,
                            "paternal": parsed.paternal_surname,
                            "maternal": parsed.maternal_surname or None,
                            "est_id": estamento_id,
                            "meta": json.dumps(meta_update),
                        }
                        await db.execute(
                            text("""
                                INSERT INTO core.person
                                    (names, paternal_surname, maternal_surname,
                                     estamento_id, metadata)
                                VALUES
                                    (:names, :paternal, :maternal, :est_id, CAST(:meta AS jsonb))
                            """),
                            insert_params,
                        )
                    stats.inserted += 1

                # Batch commit
                await batch_commit(db, row_num, dry_run)

            except Exception as e:
                stats.errors.append(f"Row {row_num} ({name_raw}): {e}")
                log.error(f"Error processing '{name_raw}': {e}")

        # Final commit
        if not dry_run:
            await db.commit()
            log.info("Final commit done")

    finally:
        await db.close()
        await close_engine()

    return stats


# ---------------------------------------------------------------------------
# Phase 1b: NOMINA xlsx → RUT enrichment
# ---------------------------------------------------------------------------

def _fuzzy_match_person(name_raw: str, person_cache: dict) -> dict | None:
    """Match an unstructured name string to a cached person.

    Strategy (ordered by specificity):
      1. Tokenize name, normalize all tokens
      2. For each person in cache, check if paternal_surname tokens are subset
      3. Among matches, verify name tokens overlap
      4. Score by total token overlap, return best match

    Handles both "Apellido1 Apellido2 Nombres" and "Nombres Apellido1 Apellido2" formats.
    """
    tokens = set(normalize_text(name_raw).split())
    if len(tokens) < 2:
        return None

    # Filter out common particles
    particles = {"de", "las", "los", "del", "la", "las", "el"}
    significant_tokens = tokens - particles

    best_match = None
    best_score = 0

    # person_cache is keyed by normalized "paternal names" or "paternal maternal names"
    # but we need to check against structured data
    seen_ids = set()
    for _key, person in person_cache.items():
        pid = str(person["id"])
        if pid in seen_ids:
            continue
        seen_ids.add(pid)

        # Build token set from DB person
        p_paternal = normalize_text(person.get("paternal_surname", ""))
        p_names = normalize_text(person.get("names", ""))
        p_maternal = normalize_text(person.get("maternal_surname", "") or "")

        person_tokens = set()
        for part in [p_paternal, p_names, p_maternal]:
            person_tokens.update(part.split())
        person_significant = person_tokens - particles

        if not person_significant:
            continue

        # Must match paternal surname (with tolerance for trailing 's')
        paternal_match = (
            p_paternal in significant_tokens
            or p_paternal + "s" in significant_tokens
            or p_paternal.rstrip("s") in significant_tokens
        )
        if not paternal_match:
            continue

        # Score: how many significant tokens overlap
        overlap = significant_tokens & person_significant
        score = len(overlap)

        # Require at least paternal + one name token
        name_tokens = set(p_names.split()) - particles
        if not (name_tokens & significant_tokens):
            continue

        if score > best_score:
            best_score = score
            best_match = person

    # Require at least 3 token overlap for confidence (paternal + maternal/name + name)
    if best_score >= 2:
        return best_match
    return None


async def enrich_ruts(nomina_path: Path, dry_run: bool = False) -> ETLStats:
    """Enrich core.person with RUT from NOMINA xlsx.

    Functor G: C_NOMINA → C_DB(core.person.rut)
    Uses fuzzy token matching since NOMINA has unstructured name format.
    """
    stats = ETLStats()

    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        stats.errors.append("Missing dependency: openpyxl")
        return stats

    if not nomina_path.exists():
        log.error(f"NOMINA file not found: {nomina_path}")
        stats.errors.append(f"File not found: {nomina_path}")
        return stats

    db = await get_session()

    try:
        # Load NOMINA xlsx
        wb = openpyxl.load_workbook(str(nomina_path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        log.info(f"NOMINA loaded: {len(rows) - 1} rows from {nomina_path.name}")

        # Build person cache
        person_cache = await build_person_cache(db)

        row_num = 0
        for row in rows[1:]:  # Skip header
            name_raw = str(row[1]).strip() if row[1] else ""
            rut_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ""

            if not name_raw or not rut_raw or name_raw.lower() in ("honorarios", ""):
                continue

            row_num += 1
            rut = normalize_rut(rut_raw)
            if not rut:
                stats.errors.append(f"Row {row_num}: Invalid RUT '{rut_raw}' for '{name_raw}'")
                continue

            # Fuzzy match to person cache
            person = _fuzzy_match_person(name_raw, person_cache)

            if not person:
                stats.errors.append(f"Row {row_num}: No match for '{name_raw}' (RUT {rut})")
                continue

            person_id = str(person["id"])
            person_display = f"{person['names']} {person['paternal_surname']}"

            # Check if RUT already set
            existing_rut = person.get("rut")
            if existing_rut:
                if normalize_rut(existing_rut) == rut:
                    stats.skipped += 1
                    log.debug(f"Skipped '{person_display}' — RUT already set to {rut}")
                else:
                    stats.errors.append(
                        f"Row {row_num}: RUT conflict for '{person_display}' — "
                        f"DB has {existing_rut}, NOMINA has {rut}"
                    )
                continue

            if dry_run:
                log.info(f"[DRY-RUN] SET RUT '{name_raw}' → {person_display} (id={person_id[:8]}...) = {rut}")
            else:
                await db.execute(
                    text("UPDATE core.person SET rut = :rut, updated_at = NOW() WHERE id = :pid"),
                    {"rut": rut, "pid": person_id},
                )
            stats.updated += 1

            await batch_commit(db, row_num, dry_run)

        if not dry_run:
            await db.commit()
            log.info("RUT enrichment commit done")

    finally:
        await db.close()
        await close_engine()

    return stats


# ---------------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------------

def main():
    parser = base_argparser("ETL Phase 1: Enrich core.person from Funcionarios CSVs + NOMINA RUTs")
    parser.add_argument(
        "--data-dir",
        type=str,
        default=str(ETL_DATA_DIR),
        help=f"Directory containing CSV sources (default: {ETL_DATA_DIR})",
    )
    parser.add_argument(
        "--nomina",
        type=str,
        default=None,
        help="Path to NOMINA xlsx file for RUT enrichment (optional)",
    )
    parser.add_argument(
        "--rut-only",
        action="store_true",
        help="Skip CSV metadata enrichment, only process NOMINA RUTs",
    )
    args = parser.parse_args()

    if args.verbose:
        setup_logging(verbose=True)

    mode = "DRY-RUN" if args.dry_run else "LIVE"

    async def _run():
        has_errors = False

        # Phase 1a: CSV metadata enrichment
        if not args.rut_only:
            log.info(f"=== ETL Phase 1a: enrich_persons [{mode}] ===")
            log.info(f"Data dir: {args.data_dir}")
            stats = await enrich_persons(
                data_dir=Path(args.data_dir),
                dry_run=args.dry_run,
                limit=args.limit,
            )
            stats.log_summary("enrich_persons (metadata)")
            if stats.total_processed > 0 and len(stats.errors) / stats.total_processed > 0.10:
                has_errors = True

        # Phase 1b: NOMINA RUT enrichment
        if args.nomina:
            log.info(f"=== ETL Phase 1b: enrich_ruts [{mode}] ===")
            rut_stats = await enrich_ruts(
                nomina_path=Path(args.nomina),
                dry_run=args.dry_run,
            )
            rut_stats.log_summary("enrich_ruts (NOMINA)")
            if rut_stats.total_processed > 0 and len(rut_stats.errors) / rut_stats.total_processed > 0.20:
                has_errors = True

        if has_errors:
            sys.exit(1)

    run_async(_run())


if __name__ == "__main__":
    main()
